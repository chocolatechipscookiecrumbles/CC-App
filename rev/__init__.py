# --- Standard Libraries ---
import os
import re
import sys
import threading
from typing import List, Tuple

# --- Third-Party Libraries ---
import pandas as pd
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
from datetime import datetime

# --- GUI (Tkinter) ---
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# --- Internal Revenue Modules ---
from .Data_Processing import (
    build_gender_sport_summaries,
    extract_unique_sports_from_rows,
    combine_financial_columns,
    normalize_sport_name,
    fix_misattributions_coeffs
)
from .Excel_Generator import (
    build_total_revenue_sheet,
    generate_template_excel_revenue
)
from .Folder_Parser import collect_revenue_across_pdfs
from .fill_excel_with_data_revenue import fill_excel_with_data_revenue
from .Testextract import extract_summary_totals
from .ui import run_ui
from .choose_uni import choose_university
from .include_client import yes_no_popup

# --- Public API ---
__all__ = [
    # Standard libs
    "os", "re", "sys", "threading", "List", "Tuple",

    # Third-party
    "pd", "pdfplumber",
    "load_workbook", "Workbook", "Font", "PatternFill", "Alignment",
    "Border", "Side", "get_column_letter", "column_index_from_string", "MergedCell",

    # Tkinter
    "tk", "filedialog", "messagebox", "ttk",

    # Internal modules
    "build_gender_sport_summaries", "extract_unique_sports_from_rows", "combine_financial_columns",
    "normalize_sport_name", "fix_misattributions_coeffs",
    "build_total_revenue_sheet", "generate_template_excel_revenue",
    "collect_revenue_across_pdfs", "fill_excel_with_data_revenue", "extract_summary_totals",
    "run_ui", "choose_university", "yes_no_popup"
]