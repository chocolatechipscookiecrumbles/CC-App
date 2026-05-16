import csv

from openpyxl import Workbook, load_workbook
import pandas as pd
import pytest

from programlauncher.common.logging_config import log_summary, start_run_log
from programlauncher.common.report_log import write_report
from programlauncher.common.run_summary import SkippedFileRecord, WorkflowRunSummary
from programlauncher.common import settings, sports
from programlauncher.common.settings_dialog import (
    AliasRowState,
    SportOpsTableRowState,
    collect_alias_settings,
    collect_sportops_table_settings,
    dedupe_aliases,
    format_aliases,
    parse_alias_text,
)


def test_run_summary_tracks_processed_skipped_and_duration():
    summary = WorkflowRunSummary("Revenue", "/tmp/pdfs", pdfs_found=2)
    summary.add_processed()
    summary.add_skipped(
        "missing.pdf",
        institution_name="Example University",
        reason="section_not_found",
        stage="revenue_extraction",
    )
    summary.finish(output_path="/tmp/revenue.xlsx")

    assert summary.processed_count == 1
    assert summary.skipped_count == 1
    assert summary.reason_counts() == {"section_not_found": 1}
    assert summary.duration_seconds is not None
    assert summary.output_path == "/tmp/revenue.xlsx"


def test_write_report_outputs_structured_csv(tmp_path):
    record = SkippedFileRecord(
        filename="bad.pdf",
        workflow="TOE",
        reason="no_extractable_text",
        stage="quality_check",
        path="/tmp/bad.pdf",
    )

    report_path = write_report(str(tmp_path), [record], workflow="TOE")

    with open(report_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    assert rows[0]["workflow"] == "TOE"
    assert rows[0]["filename"] == "bad.pdf"
    assert rows[0]["reason"] == "no_extractable_text"
    assert rows[0]["stage"] == "quality_check"
    assert rows[0]["suggested_action"]


def test_settings_round_trip_in_config_directory(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "_settings_dir", lambda: tmp_path)

    settings.remember_pdf_folder("/tmp/pdfs")
    settings.remember_save_path("/tmp/output/report.xlsx")
    settings.remember_include_client(False)

    data = settings.load_settings()
    assert data["last_pdf_folder"] == "/tmp/pdfs"
    assert data["last_save_directory"] == "/tmp/output"
    assert data["last_include_client"] is False
    assert data["recent_pdf_folders"] == ["/tmp/pdfs"]


def test_settings_store_handles_corrupt_json(tmp_path):
    path = tmp_path / "settings.json"
    path.write_text("{not json", encoding="utf-8")

    store = settings.SettingsStore(path)
    data = store.load()

    assert data["version"] == settings.SETTINGS_VERSION
    assert data["last_include_client"] is True


def test_custom_sport_aliases_extend_normalization(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "_settings_dir", lambda: tmp_path)
    settings.update_setting("custom_sport_aliases", {"Beach VB": ["Sand Volleyball"]})

    assert sports.normalize_sport_name("Sand Volleyball") == "Beach VB"


def test_authoritative_sport_aliases_override_defaults(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "_settings_dir", lambda: tmp_path)
    settings.update_setting("sport_aliases", {"Beach VB": ["Sand Volleyball"]})

    assert settings.get_sport_aliases(sports.DEFAULT_SPORT_ALIASES) == {"Beach VB": ["Sand Volleyball"]}
    assert sports.normalize_sport_name("Sand Volleyball") == "Beach VB"
    assert sports.normalize_sport_name("TrackandField") == "Trackandfield"


def test_sport_aliases_merge_legacy_custom_aliases_when_no_table_exists(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "_settings_dir", lambda: tmp_path)
    settings.update_setting("custom_sport_aliases", {"XC/TF": ["Track"], "Beach VB": ["Sand Volleyball"]})

    aliases = settings.get_sport_aliases(sports.DEFAULT_SPORT_ALIASES)

    assert aliases["XC/TF"] == [
        "Track and Field",
        "TrackandField",
        "Track and Field, X-Country",
        "Cross Country",
        "XC",
        "Track",
    ]
    assert aliases["Beach VB"] == ["Sand Volleyball"]


def test_settings_dialog_alias_text_helpers():
    assert parse_alias_text("Track and Field, TrackandField,  XC, xc, ") == [
        "Track and Field",
        "TrackandField",
        "XC",
    ]
    assert format_aliases(["Swimming and", "Swimming & Diving"]) == "Swimming and, Swimming & Diving"
    assert dedupe_aliases(["", " Soccer ", "soccer", "Football"]) == ["Soccer", "Football"]


def test_collect_alias_settings_tracks_editable_and_deleted_rows():
    class Var:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    rows = [AliasRowState(Var("XC/TF"), Var("TrackandField, TrackandField, XC"))]
    aliases = collect_alias_settings(rows)
    assert aliases == {"XC/TF": ["TrackandField", "XC"]}

    rows.append(AliasRowState(Var("test"), Var("test, test")))
    assert collect_alias_settings(rows)["test"] == ["test"]

    rows[-1].deleted = True
    assert "test" not in collect_alias_settings(rows)


def test_collect_alias_settings_cleans_empty_and_merges_duplicate_rows():
    class Var:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    rows = [
        AliasRowState(Var("Soccer"), Var("Soccer, football")),
        AliasRowState(Var(""), Var("Ignored")),
        AliasRowState(Var("Soccer"), Var("FOOTBALL, Futbol")),
        AliasRowState(Var("Empty"), Var("")),
    ]

    assert collect_alias_settings(rows) == {"Soccer": ["Soccer", "football", "Futbol"]}


def test_default_and_reset_log_directory(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "_settings_dir", lambda: tmp_path)
    assert settings.logs_dir() == tmp_path / "logs"

    settings.update_setting("log_directory", "/tmp/custom-logs")
    assert str(settings.logs_dir()) == "/tmp/custom-logs"

    settings.reset_log_directory()
    assert settings.logs_dir() == tmp_path / "logs"


def test_sportops_output_table_settings_fallback_and_reset(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "_settings_dir", lambda: tmp_path)
    default_ids = ["21", "27", "30"]

    assert settings.get_sportops_output_tables(default_ids) == default_ids

    settings.update_setting("sportops_output_tables", ["27", "999", "27"])
    assert settings.get_sportops_output_tables(default_ids) == ["27"]

    settings.update_setting("sportops_output_tables", [])
    assert settings.get_sportops_output_tables(default_ids) == default_ids

    settings.update_setting("sportops_output_tables", ["30"])
    settings.reset_sportops_output_tables(default_ids)
    assert settings.get_sportops_output_tables(default_ids) == default_ids


def test_sportops_table_settings_collect_enabled_and_protected_rows():
    class Var:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    rows = [
        SportOpsTableRowState(Var("21"), Var("Guarantees"), Var(False), is_default=True),
        SportOpsTableRowState(Var("30"), Var("Game Expenses"), Var(True), is_default=True),
        SportOpsTableRowState(Var("50"), Var("Custom Table"), Var(True)),
    ]

    table_map, selected_ids = collect_sportops_table_settings(rows)

    assert table_map == {
        "21": "Guarantees",
        "30": "Game Expenses",
        "50": "Custom Table",
    }
    assert selected_ids == ["30", "50"]

    rows[-1].deleted = True
    table_map, selected_ids = collect_sportops_table_settings(rows)
    assert "50" not in table_map
    assert selected_ids == ["30"]


def test_sportops_table_settings_clean_empty_and_duplicate_rows():
    class Var:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    rows = [
        SportOpsTableRowState(Var("21"), Var("Guarantees"), Var(True), is_default=True),
        SportOpsTableRowState(Var(""), Var("Ignored"), Var(True)),
        SportOpsTableRowState(Var("21"), Var("Duplicate"), Var(True)),
        SportOpsTableRowState(Var("88"), Var(""), Var(True)),
    ]

    table_map, selected_ids = collect_sportops_table_settings(rows)

    assert table_map == {"21": "Guarantees"}
    assert selected_ids == ["21"]


def test_sportops_table_dictionary_generates_legacy_lists():
    pytest.importorskip("pdfplumber")
    from programlauncher.sportops.config import SPORTOPS_TABLES, sportops_table_names

    assert SPORTOPS_TABLES["21"] == "Guarantees"
    assert sportops_table_names(["21", "30"]) == ["21 Guarantees", "30 Game Expenses"]


def test_sportops_workbook_headers_follow_selected_tables():
    pytest.importorskip("pdfplumber")
    from programlauncher.sportops.Excel_Generator import build_totalsports_sheet
    from programlauncher.sportops.config import sportops_table_labels

    wb = Workbook()
    ws = wb.active
    build_totalsports_sheet(
        ws,
        "CONFERENCE - Soccer Sport Operating Budgets",
        1,
        True,
        table_labels=sportops_table_labels(["21", "30"]),
    )

    assert ws["C3"].value == "Guarantees"
    assert ws["D3"].value == "Game Expenses"
    assert ws["E3"].value == "Total Operating Expenses"
    assert ws["E4"].value == "=SUM(C4:D4)"


def test_sportops_fill_filters_to_selected_tables(tmp_path):
    pytest.importorskip("pdfplumber")
    from programlauncher.sportops.Excel_Generator import generate_template_excel_totalsports
    from programlauncher.sportops.fill_excel_with_data_sportops import fill_excel_with_data_sportops

    workbook_path = tmp_path / "sportops.xlsx"
    generate_template_excel_totalsports(workbook_path)
    df = pd.DataFrame(
        {"21": [100, 10], "30": [200, 20], "40": [999, 99]},
        index=["Client University", "Peer University"],
    )

    fill_excel_with_data_sportops(
        {"Men's Soccer": df},
        workbook_path,
        "Client University",
        ["Men's Soccer"],
        [],
        True,
        table_ids=["21", "30"],
    )

    loaded = load_workbook(workbook_path, data_only=False)
    ws = loaded["Men's Soccer"]
    assert ws["C3"].value == "Guarantees"
    assert ws["D3"].value == "Game Expenses"
    assert ws["E3"].value == "Total Operating Expenses"
    assert ws["C4"].value == 100
    assert ws["D4"].value == 200
    assert ws["E4"].value == "=SUM(C4:D4)"


def test_revenue_validation_sheet_marks_ticket_sales_status(tmp_path):
    pytest.importorskip("pdfplumber")
    from programlauncher.rev.fill_excel_with_data_revenue import write_revenue_validation_sheet

    workbook_path = tmp_path / "revenue.xlsx"
    wb = Workbook()
    write_revenue_validation_sheet(
        wb,
        [
            {
                "institution_name": "Example University",
                "filename": "example.pdf",
                "ticket_sales_status": "missing",
                "ticket_sales_value": None,
                "missing_tables": ["1"],
                "found_tables": ["2"],
                "combined_categories": ["2 = 2 + 4"],
                "skipped": False,
                "parse_error": "",
            }
        ],
    )
    wb.save(workbook_path)

    loaded = load_workbook(workbook_path)
    ws = loaded["Validation"]
    assert ws["A2"].value == "Example University"
    assert ws["C2"].value == "missing"
    assert ws["E2"].value == "1"


def test_run_logging_writes_summary(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "_settings_dir", lambda: tmp_path)
    summary = WorkflowRunSummary("Revenue", "/tmp/pdfs", pdfs_found=1)
    start_run_log(summary)
    summary.add_processed()
    summary.finish(output_path="/tmp/revenue.xlsx")
    log_summary(summary)

    log_path = summary.extra["log_path"]
    assert "Revenue" in open(log_path, encoding="utf-8").read()
