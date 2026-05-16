from openpyxl import Workbook

from scripts.compare_workbooks import build_report, compare_workbooks


def _save_workbook(path, value):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = value
    ws["B1"] = "=SUM(1,2)"
    wb.save(path)


def test_compare_workbooks_identical_files_are_clean(tmp_path):
    baseline = tmp_path / "baseline.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _save_workbook(baseline, "same")
    _save_workbook(candidate, "same")

    findings = compare_workbooks(str(baseline), str(candidate))
    report = build_report(findings, str(baseline), str(candidate))

    assert findings == []
    assert "No workbook differences found." in report


def test_compare_workbooks_reports_value_changes(tmp_path):
    baseline = tmp_path / "baseline.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    _save_workbook(baseline, "old")
    _save_workbook(candidate, "new")

    findings = compare_workbooks(str(baseline), str(candidate))

    assert any("A1" in finding for finding in findings)
