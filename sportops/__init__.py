from __future__ import annotations
# --- Standard Libraries ---
import os
import re
import sys
import threading


# --- Third-Party Libraries ---
import pandas as pd
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
from datetime import datetime

# --- GUI (Tkinter) ---
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# --- Internal Sport Ops Modules ---
from .Data_Processing import (
    build_gender_sport_summaries,
    extract_unique_sports_from_rows,
    normalize_sport_name,
    fix_misattributions_coeffs
)
from .Excel_Generator import (
    build_totalsports_sheet,
    build_total_sports_total_sheet,
    generate_template_excel_totalsports
)
from .Folder_Parser import collect_sports_across_pdfs
from .fill_excel_with_data_sportops import fill_excel_with_data_sportops
from .Table_Extractor import extract_tables_by_title
from .ui import run_ui
from .choose_uni import choose_university
from .checklist import generate_checklist
from .include_client import yes_no_popup

# --- Public API ---
__all__ = [
    # Standard libs
    "os", "re", "sys", "threading",

    # Third-party
    "pd", "pdfplumber",
    "load_workbook", "Workbook", "Font", "PatternFill", "Alignment",
    "Border", "Side", "get_column_letter", "column_index_from_string", "MergedCell",

    # Tkinter
    "tk", "filedialog", "messagebox", "ttk",

    # Internal modules
    "build_gender_sport_summaries", "extract_unique_sports_from_rows",
    "normalize_sport_name", "fix_misattributions_coeffs",
    "build_totalsports_sheet", "build_total_sports_total_sheet",
    "generate_template_excel_totalsports",
    "collect_sports_across_pdfs", "fill_excel_with_data_sportops",
    "extract_tables_by_title",
    "run_ui", "choose_university", "generate_checklist", "yes_no_popup"
]