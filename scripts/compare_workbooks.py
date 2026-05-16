"""Compare two generated Excel workbooks and write a Markdown report."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def _cell_value(cell):
    return cell.value


def compare_workbooks(baseline_path: str, candidate_path: str) -> list[str]:
    baseline = load_workbook(baseline_path, data_only=False)
    candidate = load_workbook(candidate_path, data_only=False)

    findings: list[str] = []
    baseline_sheets = baseline.sheetnames
    candidate_sheets = candidate.sheetnames

    if baseline_sheets != candidate_sheets:
        findings.append(
            f"Sheet order changed: baseline={baseline_sheets}, candidate={candidate_sheets}"
        )

    missing_sheets = [name for name in baseline_sheets if name not in candidate_sheets]
    extra_sheets = [name for name in candidate_sheets if name not in baseline_sheets]
    for name in missing_sheets:
        findings.append(f"Missing sheet in candidate: `{name}`")
    for name in extra_sheets:
        findings.append(f"Extra sheet in candidate: `{name}`")

    for sheet_name in [name for name in baseline_sheets if name in candidate_sheets]:
        base_ws = baseline[sheet_name]
        cand_ws = candidate[sheet_name]

        if base_ws.max_row != cand_ws.max_row:
            findings.append(
                f"`{sheet_name}` row count changed: {base_ws.max_row} -> {cand_ws.max_row}"
            )
        if base_ws.max_column != cand_ws.max_column:
            findings.append(
                f"`{sheet_name}` column count changed: {base_ws.max_column} -> {cand_ws.max_column}"
            )

        max_row = max(base_ws.max_row, cand_ws.max_row)
        max_col = max(base_ws.max_column, cand_ws.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                base_cell = base_ws.cell(row=row, column=col)
                cand_cell = cand_ws.cell(row=row, column=col)
                base_value = _cell_value(base_cell)
                cand_value = _cell_value(cand_cell)
                if base_value != cand_value:
                    findings.append(
                        f"`{sheet_name}` {base_cell.coordinate}: {base_value!r} -> {cand_value!r}"
                    )

    return findings


def build_report(findings: Iterable[str], baseline_path: str, candidate_path: str) -> str:
    findings = list(findings)
    lines = [
        "# Workbook Comparison Report",
        "",
        f"- Baseline: `{baseline_path}`",
        f"- Candidate: `{candidate_path}`",
        "",
    ]
    if not findings:
        lines.extend(["## Result", "", "No workbook differences found."])
    else:
        lines.extend(["## Differences", ""])
        lines.extend(f"- {finding}" for finding in findings)
    lines.append("")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Compare two Excel workbooks.")
    parser.add_argument("baseline", help="Path to the baseline workbook")
    parser.add_argument("candidate", help="Path to the candidate workbook")
    parser.add_argument(
        "--output",
        default="comparison_report.md",
        help="Markdown report output path",
    )
    args = parser.parse_args(argv)

    findings = compare_workbooks(args.baseline, args.candidate)
    report = build_report(findings, args.baseline, args.candidate)
    Path(args.output).write_text(report, encoding="utf-8")
    print(f"Wrote comparison report to {args.output}")
    return 1 if findings else 0


if __name__ == "__main__":
    raise SystemExit(main())
