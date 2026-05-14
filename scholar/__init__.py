# --- Standard Libraries ---
import os
import re
import sys
import threading

# --- Third-Party Libraries ---
import pandas as pd
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from collections import Counter
from datetime import datetime

# --- GUI (Tkinter) ---
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# --- Internal Scholar Modules ---
from .ui import run_ui
from .build_scholarship_sheet import generate_template_excel_scholar
from .extract_scholarships import process_folder
from .fill_excel_with_data_scholar import fill_excel_with_data_scholar
from .checklist import generate_checklist
from .choose_uni import choose_university
from .include_client import yes_no_popup

# --- Public API ---
__all__ = [
    # Standard libs
    "os", "re", "sys", "threading",
    # Third-party
    "pd", "pdfplumber",
    "load_workbook", "Workbook", "Font", "PatternFill", "Alignment",
    "Border", "Side", "get_column_letter", "MergedCell",
    # Tkinter
    "tk", "filedialog", "messagebox", "ttk",
    # Internal
    "run_ui", "generate_template_excel_scholar", "process_folder",
    "fill_excel_with_data_scholar", "generate_checklist",
    "choose_university", "yes_no_popup",
]